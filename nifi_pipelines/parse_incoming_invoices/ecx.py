import os
import sys
import json
import requests
from datetime import datetime
import openpyxl
from io import BytesIO
from dotenv import load_dotenv
import uuid

# Load environment variables
load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
FOLDER_ID = os.getenv("ONEDRIVE_FOLDER_ID")

# Microsoft Graph API endpoints
AUTH_URL = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
GRAPH_API_URL = "https://graph.microsoft.com/v1.0"

# File naming for May 2025
MONTH_YEAR = datetime.now().strftime("%b_%Y").upper()
MANUAL_FILE = f"Invoices_{MONTH_YEAR}.xlsx"
BACKUP_FILE = f"Invoices_BACKUP_{MONTH_YEAR}.xlsx"
PIPELINE_FILE = f"Invoices_PIPELINE_{MONTH_YEAR}.xlsx"

# Excel schema
SCHEMA = [
    "Vendor Name", "System Processing Date", "Invoice Number", "Description",
    "Invoice Date", "Total Amount", "GST Amount", "TDS", "Amount Payable", "Created By"
]

def get_access_token():
    """Obtain access token using client credentials."""
    payload = {
        "client_id": CLIENT_ID,
        "scope": "https://graph.microsoft.com/.default",
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials"
    }
    response = requests.post(AUTH_URL, data=payload)
    response.raise_for_status()
    return response.json()["access_token"]

def get_file_id(folder_id, file_name, headers):
    """Search for a file in the OneDrive folder by name."""
    url = f"{GRAPH_API_URL}/drives/{folder_id}/items/root/children"
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    files = response.json().get("value", [])
    for file in files:
        if file["name"] == file_name:
            return file["id"]
    return None

def download_excel(file_id, headers):
    """Download Excel file content from OneDrive."""
    url = f"{GRAPH_API_URL}/drives/{folder_id}/items/{file_id}/content"
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return BytesIO(response.content)

def upload_excel(file_id, content, headers):
    """Upload Excel file content to OneDrive."""
    url = f"{GRAPH_API_URL}/drives/{folder_id}/items/{file_id}/content"
    response = requests.put(url, headers=headers, data=content)
    response.raise_for_status()

def create_excel(folder_id, file_name, headers):
    """Create a new Excel file with the specified schema."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(SCHEMA)
    content = BytesIO()
    wb.save(content)
    content.seek(0)
    url = f"{GRAPH_API_URL}/drives/{folder_id}/items/root:/{file_name}:/content"
    response = requests.put(url, headers=headers, data=content)
    response.raise_for_status()
    return response.json()["id"]

def are_sheets_identical(backup_wb, manual_wb):
    """Compare two Excel workbooks for identical content."""
    backup_ws = backup_wb.active
    manual_ws = manual_wb.active
    if backup_ws.max_row != manual_ws.max_row or backup_ws.max_column != manual_ws.max_column:
        return False
    for row1, row2 in zip(backup_ws.iter_rows(values_only=True), manual_ws.iter_rows(values_only=True)):
        if row1 != row2:
            return False
    return True

def get_additional_rows(manual_wb, backup_wb):
    """Fetch additional rows from manual sheet not present in backup."""
    manual_ws = manual_wb.active
    backup_ws = backup_wb.active
    additional_rows = []
    for row in manual_ws.iter_rows(min_row=2, values_only=True):
        if row not in backup_ws.iter_rows(min_row=2, values_only=True):
            additional_rows.append(row)
    return additional_rows

def main():
    # Read JSON input from stdin
    invoice_data = json.load(sys.stdin)
    
    # Get access token
    headers = {"Authorization": f"Bearer {get_access_token()}"}
    
    # Get or create Excel files
    manual_file_id = get_file_id(FOLDER_ID, MANUAL_FILE, headers)
    backup_file_id = get_file_id(FOLDER_ID, BACKUP_FILE, headers)
    pipeline_file_id = get_file_id(FOLDER_ID, PIPELINE_FILE, headers)
    
    # If manual file doesn't exist, create all three files
    if not manual_file_id:
        manual_file_id = create_excel(FOLDER_ID, MANUAL_FILE, headers)
        backup_file_id = create_excel(FOLDER_ID, BACKUP_FILE, headers)
        pipeline_file_id = create_excel(FOLDER_ID, PIPELINE_FILE, headers)
    
    # Download and create backup
    manual_wb = openpyxl.load_workbook(download_excel(manual_file_id, headers))
    backup_content = BytesIO()
    manual_wb.save(backup_content)
    if not backup_file_id:
        backup_file_id = create_excel(FOLDER_ID, BACKUP_FILE, headers)
    else:
        upload_excel(backup_file_id, backup_content, headers)
    
    # Load or create pipeline sheet
    if pipeline_file_id:
        pipeline_wb = openpyxl.load_workbook(download_excel(pipeline_file_id, headers))
    else:
        pipeline_wb = openpyxl.Workbook()
        pipeline_wb.active.append(SCHEMA)
        pipeline_file_id = create_excel(FOLDER_ID, PIPELINE_FILE, headers)
    
    # Append invoice data to pipeline sheet
    pipeline_ws = pipeline_wb.active
    today = datetime.now().strftime("%Y-%m-%d")
    for invoice in invoice_data:
        pipeline_ws.append([
            invoice.get("VendorName"),
            today,
            str(uuid.uuid4())[:8],  # Generate unique invoice number
            "",  # Description (not provided in input)
            invoice.get("Date"),
            invoice.get("Amount_Payable"),
            invoice.get("GST"),
            invoice.get("TDS"),
            invoice.get("Amount_Payable"),
            "Pipeline"  # Created By
        ])
    
    # Save pipeline sheet
    pipeline_content = BytesIO()
    pipeline_wb.save(pipeline_content)
    upload_excel(pipeline_file_id, pipeline_content, headers)
    
    # Compare backup and manual sheets
    backup_wb = openpyxl.load_workbook(download_excel(backup_file_id, headers))
    manual_wb = openpyxl.load_workbook(download_excel(manual_file_id, headers))
    
    if are_sheets_identical(backup_wb, manual_wb):
        # If identical, update manual sheet with pipeline data
        upload_excel(manual_file_id, pipeline_content, headers)
    else:
        # Fetch additional rows from manual sheet and append to pipeline
        additional_rows = get_additional_rows(manual_wb, backup_wb)
        for row in additional_rows:
            pipeline_ws.append(row)
        pipeline_content = BytesIO()
        pipeline_wb.save(pipeline_content)
        upload_excel(pipeline_file_id, pipeline_content, headers)

if __name__ == "__main__":
    main()